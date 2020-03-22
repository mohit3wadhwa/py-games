import pgzrun
from random import randint
import openpyxl
import time
'''
KNOW YOUR PARTNER GAME WITH EXCEL AS DB
'''

WIDTH  = 1280
HEIGHT = 720


title_box   =  Rect(0, 0, 1280, 30)
main_box    =  Rect(0, 0, 850, 200)
timer_box   =  Rect(0, 0, 240, 200)
answer_box1 =  Rect(0, 0, 550, 165)
answer_box2 =  Rect(0, 0, 550, 165)
answer_box3 =  Rect(0, 0, 550, 165)
answer_box4 =  Rect(0, 0, 550, 165)


title_box.move_ip(0, 10)
main_box.move_ip(50, 70)
timer_box.move_ip(990, 70)
answer_box1.move_ip(50, 358)
answer_box2.move_ip(685, 358)
answer_box3.move_ip(50, 538)
answer_box4.move_ip(685, 538)

# Create Sheet Object & get the sheet name
wb = openpyxl.load_workbook('Q_A.xlsx')
sheet_obj = wb.get_sheet_by_name('Sheet1')

# Get the total number of rows 
total_xls_enrtries = sheet_obj.max_row

questions = []
options = []
correct_answers = []
option_starts = 4


# Loop will print all values  
for i in range(2, total_xls_enrtries + 1):
    cell_obj1 = sheet_obj.cell(row = i, column = 2)
    questions.append(cell_obj1.value)  

    for j in range(option_starts, 9):
        cell_obj2 = sheet_obj.cell(row = i, column = j)
        if j == 8:
            correct_answers.append(cell_obj2.value)
        else:
            options.append(cell_obj2.value)

question = questions.pop(0)

answer_boxes = [answer_box1, answer_box2, answer_box3, answer_box4]
score = 0
time_left = 5
color_var = 'orange'

def draw():
    screen.fill("dim grey")
    screen.draw.filled_rect(title_box, "pink")
    screen.draw.filled_rect(main_box, "sky blue")
    screen.draw.filled_rect(timer_box, "sky blue")

    for box in answer_boxes:
        screen.draw.filled_rect(box, color_var)

    screen.draw.textbox(str(time_left), timer_box, color=("black"))
    screen.draw.textbox("L E T ' S   P L A Y  ,  K N O W   Y O U R   P A R T N E R . . .", title_box, color=("black"))
    screen.draw.textbox(question, main_box, color=("black"))

    index  = 0
    for box in answer_boxes:
        screen.draw.textbox(str(options[index]), box, color=("black"))
        index += 1

def game_over(var):
    #global game_over_flag
    global time_left
    global question
    question = "Game Over! Your Final score is: " + str(score)
    correct_answers[0] = 5
    for i in range(0, 4):
        options[i] = "-"
    time_left = 0


def correct_answer():
    global score
    global time_left
    global question
  
    score = score + 1
    if questions:
        question = questions.pop(0)
        correct_answers.pop(0)
        for i in range(1, 5):
            options.pop(0)
        time_left = 5
    else:
        game_over("N")

def on_mouse_down(pos):
    global color_var
    index = 1
    #global correct_answers
    for box in answer_boxes:
        if box.collidepoint(pos):
            if index == correct_answers[0]:
                correct_answer()
            else:
                game_over("Y")
        index += 1


def updates_time_left():
    global time_left
    if time_left != 0:
        time_left -= 1
    else:
       game_over("Y")

clock.schedule_interval(update_time_left, 1.0)
pgzrun.go()
